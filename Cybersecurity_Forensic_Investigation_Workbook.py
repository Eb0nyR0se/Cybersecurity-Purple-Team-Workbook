import pandas as pd
import os
from openpyxl.styles import Alignment, Font, PatternFill

# Path Setup
desktop_path = os.path.join(os.environ['USERPROFILE'], 'OneDrive', 'Desktop')
file_name = os.path.join(desktop_path, "Cybersecurity_Forensic_Investigation_Workbook.xlsx")

# --- DATA STRUCTURES (Balanced for Tier-3 Standards) ---

# 1. TACTICAL LOG: The "Command & Output" Workbench
# This tracks specific tools, exact commands, and raw terminal findings.
tactical_log = {
    "Timestamp": ["10:00", "10:15", "10:30", "10:45", "11:00", "11:15"],
    "Platform/OS": ["Kali Purple", "Security Onion", "Metasploit", "Wireshark", "PowerShell", "Autopsy"],
    "Tool Category": ["Network Recon", "Intrusion Detection", "Exploitation", "Traffic Analysis", "Host Intel",
                      "Digital Forensics"],
    "Command Executed": [
        "nmap -sV -p- 192.168.x.x",
        "so-status / zeek-logs",
        "use exploit/multi/handler",
        "tcpdump -i eth0 -w dump.pcap",
        "Get-NetIPConfiguration",
        "Analyze E01 Disk Image"
    ],
    "Raw Findings/Output": [
        "Port 445 (SMB) Open",
        "Zeek recording L2 traffic",
        "Payload: windows/x64/meterpreter",
        "Captured 450 packets",
        "MAC: 92:65:35 verified",
        "Deleted .txt files recovered"
    ],
    "Evidence ID": ["N/A", "EV-001", "N/A", "EV-002", "N/A", "EV-003"]
}

# 2. NETWORK PENTEST: PTES & OSSTMM (Aligned with Industry Standards)
pentest_matrix = {
    "PTES Phase": ["Pre-engagement", "Intelligence Gathering", "Vulnerability Analysis", "Exploitation",
                   "Post-Exploitation", "Reporting"],
    "OSSTMM Class": ["Physical", "Spectrum", "Wireless", "Telecommunications", "Data Networks", "Human"],
    "Primary Tools": ["Contract Docs", "theHarvester", "Nmap / Nessus", "Metasploit", "Mimikatz", "Python/Pandas"],
    "Methodology": ["Scope Definition", "DNS/OSINT Recon", "Service Mapping", "Payload Delivery",
                    "Privilege Escalation", "Final Dossier"],
    "Result/Vuln ID": ["Verified Scope", "Subdomains Found", "CVE-2023-XXXX", "User Shell", "Domain Admin",
                       "Workbook Complete"]
}

# 3. WEB APP SECURITY: OWASP Top 10 Tracker
owasp_data = {
    "OWASP Category": ["A01:2021-Broken Access Control", "A03:2021-Injection", "A07:2021-Identification/Auth Failures"],
    "Target / URL": ["https://phoenix.local/admin", "Login Query Field", "Session Cookie"],
    "Test Method": ["IDOR / Manual Manipulation", "SQLi Fuzzing", "Brute Force / Hydra"],
    "Tools Used": ["Burp Suite / Zap", "SQLmap", "Custom Python Script"],
    "Risk Level": ["Critical", "High", "Medium"],
    "Remediation": ["Implement RBAC", "Parameterized Queries", "MFA Enforcement"]
}

# 4. INCIDENT RESPONSE: NIST SP 800-61 Lifecycle
nist_ir_lifecycle = {
    "NIST Phase": ["Preparation", "Detection/Analysis", "Containment", "Eradication", "Recovery"],
    "Action Taken": ["Sensor Hardening", "Alert Correlation", "VLAN Isolation", "Malware Cleanup", "Full Restore"],
    "Tools Applied": ["pfSense / Snort", "Security Onion", "Kali Purple", "ClamAV / Autopsy", "Backup Exec"],
    "Integrity Hash": ["N/A", "SHA256: e3b0c442...", "N/A", "SHA256: 5df6e0e2...", "N/A"],
    "Status": ["Complete", "Verified", "Active", "In Progress", "Pending"]
}

# 5. EVIDENCE VAULT: Chain of Custody
evidence_vault = {
    "Evidence ID": ["EV-001", "EV-002", "EV-003"],
    "Collection Date": ["2026-05-11", "2026-05-11", "2026-05-11"],
    "Artifact Source": ["Security Onion Log", "Wireshark PCAP", "Autopsy Disk Image"],
    "SHA-256 Hash": ["e3b0c442...", "5df6e0e2...", "a665a459..."],
    "Forensic Analyst": ["First Last Name", "First Last Name", "First Last Name"],
    "Integrity Status": ["Verified", "Verified", "Verified"]
}


# --- FORMATTING ENGINE ---

def apply_elite_style(worksheet, title, col_count, theme_color):
    """Applies high-end professional formatting to each tab."""
    # Main Header
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    top_cell = worksheet.cell(row=1, column=1)
    top_cell.value = title

    # Visual Aesthetics
    fill = PatternFill(start_color=theme_color, end_color=theme_color, fill_type="solid")
    font = Font(color="FFFFFF", bold=True, size=14)
    alignment = Alignment(horizontal="center")

    top_cell.fill = fill
    top_cell.font = font
    top_cell.alignment = alignment

    # Column Header Styling (Row 2)
    for col in range(1, col_count + 1):
        cell = worksheet.cell(row=2, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = alignment


def generate_workbook():
    try:
        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            # Generate Sheets
            pd.DataFrame(tactical_log).to_excel(writer, sheet_name='Tactical_Activity_Log', index=False, startrow=1)
            pd.DataFrame(pentest_matrix).to_excel(writer, sheet_name='Network_PenTest_PTES', index=False, startrow=1)
            pd.DataFrame(owasp_data).to_excel(writer, sheet_name='Web_App_OWASP', index=False, startrow=1)
            pd.DataFrame(nist_ir_lifecycle).to_excel(writer, sheet_name='NIST_IR_Lifecycle', index=False, startrow=1)
            pd.DataFrame(evidence_vault).to_excel(writer, sheet_name='Evidence_Vault', index=False, startrow=1)

            # Applying Professional Color Themes
            apply_elite_style(writer.sheets['Tactical_Activity_Log'], "TACTICAL COMMAND & TOOL LOG", 6,
                              "4472C4")  # Blue
            apply_elite_style(writer.sheets['Network_PenTest_PTES'], "PTES & OSSTMM: NETWORK PENETRATION TESTING", 5,
                              "385723")  # Green
            apply_elite_style(writer.sheets['Web_App_OWASP'], "OWASP TOP 10: WEB APPLICATION ASSESSMENT", 6,
                              "E65100")  # Orange
            apply_elite_style(writer.sheets['NIST_IR_Lifecycle'], "NIST SP 800-61: INCIDENT RESPONSE LIFECYCLE", 5,
                              "0D47A1")  # Dark Blue
            apply_elite_style(writer.sheets['Evidence_Vault'], "DFIR: FORENSIC CHAIN OF CUSTODY & HASH VAULT", 6,
                              "212121")  # Black

        print("-" * 65)
        print(f"TIER-3 WORKBOOK GENERATED SUCCESSFULLY:\n{file_name}")
        print("-" * 65)
    except Exception as e:
        print(f"CRITICAL SYSTEM ERROR: {e}")


if __name__ == "__main__":
    generate_workbook()
