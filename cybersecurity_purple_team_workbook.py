import pandas as pd
from datetime import datetime
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.chart import (
    AreaChart,
    BarChart,
    DoughnutChart,
    RadarChart,
    PieChart,
    LineChart,
    Reference
)
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell
import os


def purple_team_workbook():
    # --- 1. ROBUST ENVIRONMENT PATHING ---
    user_profile = os.environ.get('USERPROFILE', os.path.expanduser('~'))
    paths = [
        os.path.join(user_profile, 'OneDrive', 'Desktop'),
        os.path.join(user_profile, 'Desktop'),
        user_profile
    ]
    desktop = next((p for p in paths if os.path.exists(p)), user_profile)
    if not os.path.exists(desktop):
        desktop = os.path.expanduser("~/Desktop")

    filename = os.path.join(desktop, "Cybersecurity_Unified_Operations_Workbook.xlsx")

    # Identity & Timing
    analyst_name = ""
    today_str = datetime.now().strftime('%Y-%m-%d')
    now_ts = datetime.now().strftime('%H:%M:%S')

    # Color Constants
    BLUE_TEAM_COLOR, RED_TEAM_COLOR = "012169", "8B0000"
    PURPLE_THEME, LILAC_PURPLE = "4A148C", "E1BEE7"
    DARK_GREY, NAVY_BLUE = "212121", "01579B"
    BLUE_TEAM_LIGHT, RED_TEAM_LIGHT = "BBDEFB", "FFCDD2"
    BLUE_TEAM_MID, RED_TEAM_MID = "0D47A1", "B71C1C"

    # --- 2. DATA TABLES ---

    # TAB 1: Executive Summary
    summary_dashboard = {
        'KPI_Metric': ['Total Risks', 'Open Mitigations', 'Mean Time to Detect', 'Critical Vulns', 'Engagement Status'],
        'Current_Value': ['3', '2', '12 Minutes', '4', 'Active - SOC Monitoring'],
        'Threshold': ['< 5', '0', '< 30 Minutes', '0', 'Operational']
    }
    # TAB 2: DOSSIER
    dossier_rows = [
        ['--- CASE IDENTIFICATION ---', '', ''],
        ['Case ID', f'CASE-{datetime.now().year}-001', ''],
        ['Analyst', analyst_name, ''],
        ['Engagement Type', 'Purple Team Engagement', ''],
        ['Classification', 'Internal Only', ''],
        ['', '', ''],
        ['--- EXECUTIVE SUMMARY ---', '', 'Threshold'],
        ['Total Risks', '3', '< 5'],
        ['Open Mitigations', '2', '0'],
        ['Mean Time to Detect', '12 Minutes', '< 30 Minutes'],
        ['Critical Vulns', '4', '0'],
        ['Engagement Status', 'Active - SOC Monitoring', 'Operational'],
        ['', '', ''],
        ['--- INFRASTRUCTURE ---', '', ''],
        ['Lab Standard', 'NIST / MITRE', ''],
        ['Primary Firewall', 'Netgate 2100', ''],
        ['IDS Engine', 'Security Onion', ''],
        ['Managed Switch', 'Netgear GS305E', ''],
        ['Baseline OS', 'Aegis OS', ''],
        ['Network Zone', 'Internal Lab', '']
    ]

    # TAB 3: EVIDENCE VAULT
    ev_data = {
        'Date': [today_str, today_str, today_str],
        'Timestamp': [now_ts, now_ts, now_ts],
        'Evidence_ID': ['EV-001', 'EV-002', 'EV-003'],
        'Artifact_Name': ['log4j_exploit_capture.pcap', 'Nmap Scan Log', 'Wireshark PCAP'],
        'Artifact_Source': ['Netgate 2100 / SPAN', 'Nmap Scan Log', 'Wireshark PCAP'],
        'Custodian': ['Network_Admin', 'Beth', 'Beth'],
        'Collection_Point': ['Netgate 2100 / SPAN', 'Lab Console', 'Security Onion Sensor'],
        'Collection_Tool': ['tcpdump', 'nmap', 'Wireshark'],
        'SHA-256_Hash': ['e3b0c442...', 'e3b0c442...', '5df6e0e2...'],
        'Hash_Verified': ['TRUE', 'TRUE', 'TRUE'],
        'Integrity_Status': ['Verified', 'Verified', 'Verified'],
        'Storage_Location': ['Vault_Drive_01', 'Vault_Drive_01', 'Vault_Drive_01'],
        'File_Size_MB': ['45.2', '0.1', '112.8'],
        'CVE_Ref': ['CVE-2021-44228', 'N/A', 'N/A'],
        'Status': ['Locked/Verified', 'Locked/Verified', 'Locked/Verified'],
        'Retention_Period': ['1 Year', '1 Year', '1 Year']
    }

    # NEW TAB 4: LESSONS LEARNED (Enhanced)
    lessons_learned_rows = [
        [
            'Issue Identified', 'Root Cause', 'Detection Gap',
            'MITRE Ref', 'Remediation Action', 'Complexity',
            'Risk Reduction', 'Validation Test', 'Owner', 'Status'
        ],
        [
            'Log4j alerts delayed', 'ES indexing lag', 'Process',
            'T1190', 'Optimize Indexing', 'Medium',
            'High', 'Re-run Log4Shell exploit', 'DevOps', 'Pending'
        ],
        [
            'SMBExec undetected', 'Missing Sysmon EID 1', 'Visibility',
            'T1021.001', 'Update Sysmon Config', 'Low',
            'Critical', 'Invoke-AtomicTest T1021.001', 'Blue Team', 'Open'
        ]
    ]

    fe_data = {
        'Evidence_ID': ['EV-001'],
        'Analyst_ID': [''],
        'Acquisition_Type': ['Live Network Capture'],
        'Source_Device': ['Netgate-VLAN-20-SPAN'],  # MISSING
        'Platform': ['Ubuntu 22.04'],
        'Artifact_Examined': ['log4j.pcap'],
        'Source_Hash_SHA256': ['e3b0c442...'],  # MISSING
        'Forensic_Hash_Verify': ['MATCH'],
        'Timezone': ['UTC-4'],  # MISSING
        'Tool_Used': ['Wireshark 4.0.2 / Zeek 5.0'],  # Versions added
        'Analysis_Machine': ['Forensic-WS-01'],  # MISSING
        'Findings_Summary': ['JNDI injection detected; China Chopper callback established.'],
        'Timeline_Match': ['Confirmed via Sysmon'],
        'Action_Taken': ['Isolation'],
        'Status': ['Completed'],
        'Event_Date': ['2026-05-11'],
        'Event_Time': ['14:30:00']
    }

    # TAB 5: TOPOLOGY
    topo_cols = [
        'VLAN', 'Name', 'Network', 'Gateway', 'Phys_Port', 'Interface',
        'Assignment', 'MTU', 'Traffic_Mode', 'SPAN_Dest', 'Security_Policy'
    ]
    topo_rows = [
        # VLAN | Name | Network | Gateway | Phys_Port | Interface | Assignment | MTU | Traffic_Mode | SPAN_Dest | Security_Policy
        ['VLAN 10', 'Management', '192.168.10.0/24', '192.168.10.1', 'Port 1', 'mvneta0.10', 'Static', '1500', 'Source',
         'N/A', 'Admin_Only'],
        ['VLAN 20', 'Victim_Lab', '192.168.20.0/24', '192.168.20.1', 'Port 2', 'mvneta0.20', 'DHCP', '1500', 'Mirrored',
         'Port 4', 'Isolated_No_WAN'],
        ['VLAN 30', 'Attack_Srv', '192.168.30.0/24', '192.168.30.1', 'Port 3', 'mvneta0.30', 'Static', '1500',
         'Mirrored', 'Port 4', 'Strict_Egress'],
        ['VLAN 40', 'Security_Onion', '10.0.40.0/24', '10.0.40.1', 'Port 4', 'mvneta0.40', 'Static', '9000',
         'Destination', 'N/A', 'Monitor_All'],
        ['VLAN 99', 'Monitoring', 'Mirror Port', 'Netgate 2100', 'Mirror', 'Security Onion', 'Static', '1500',
         'Destination', 'N/A', 'Monitor_All']
    ]

    # TAB 6: PURPLE STRATEGY (Restored Full Descriptions)
    ps_rows = [
        ['--- BLUE TEAM ENCYCLOPEDIA ---', '', ''],
        ['TOOL', 'FUNCTIONAL WRITEUP', 'KNOWN COMMANDS'],
        ['SECURITY ONION', 'Full NSM/IDS stack platform for network monitoring and log management.', 'so-status; so-allow; so-repro'],
        ['ZEEK', 'Metadata and protocol analysis tool for deep network visibility and flow forensics.', 'zeekctl status; zeekctl start'],
        ['SURICATA', 'Signature-based IDS/IPS engine for real-time threat detection and alerting.', 'suricata -c /etc/suricata/suricata.yaml'],
        ['AUTOPSY', 'Digital forensics GUI platform for disk image analysis and artifact recovery.', 'autopsy &'],
        ['VOLATILITY', 'Advanced memory forensics framework for analyzing RAM dumps.', 'vol.py -f mem.raw windows.info'],
        ['SYSINTERNALS', 'Windows endpoint triage suite for process, network, and persistence inspection.', 'autoruns; procmon; tcpview'],
        ['OSQUERY', 'SQL-like endpoint telemetry for hunting and compliance checks.', 'osqueryi; SELECT * FROM processes;'],
        ['', '', ''],
        ['--- BLUE TEAM STRATEGY: NIST IR FRAMEWORK ---', '', ''],
        ['PHASE', 'STRATEGIC FOCUS', 'GOAL'],
        ['PREPARATION', 'Harden systems, log baselines, and establish incident response plans.', 'Reduce attack surface.'],
        ['DETECTION', 'Alert triage, IOC correlation, and analysis of anomalous traffic.', 'Identify gaps in visibility.'],
        ['CONTAINMENT', 'Isolate affected network segments and preserve forensic evidence.', 'Stop the spread of the attack.'],
        ['ERADICATION', 'Remove malware, disable compromised accounts, and close exploited paths.', 'Ensure threat is removed.'],
        ['RECOVERY', 'Restore affected services and validate system integrity before returning to normal operations.', 'Return to business-as-usual safely.'],
        ['LESSONS LEARNED', 'Capture root causes, missed telemetry, and remediation owners.', 'Improve the next engagement cycle.'],
        ['', '', ''],
        ['--- RED TEAM ENCYCLOPEDIA ---', '', ''],
        ['TOOL', 'FUNCTIONAL WRITEUP', 'KNOWN COMMANDS'],
        ['NMAP', 'Network discovery and port scanning.', 'nmap -sS -sV -O -p- [target]'],
        ['NETDISCOVER', 'Active/passive ARP reconnaissance.', 'netdiscover -i eth0 -r 192.168.1.0/24'],
        ['METASPLOIT', 'Exploit development and execution.', 'msfconsole; search [CVE]; exploit'],
        ['MSFVENOM', 'Payload generation/encoding.', 'msfvenom -p windows/x64/reverse_tcp'],
        ['COBALT STRIKE', 'Adversary simulation/C2 beacons.', './teamserver [IP] [Pass]; ./beacon'],
        ['CHINA CHOPPER', 'Webshell for persistent access.', 'C2 client interaction via URL'],
        ['MIMIKATZ', 'Credential theft and ticket attacks.', 'privilege::debug; sekurlsa::logonpasswords'],
        ['IMPACKET', 'Python toolkit for SMB, Kerberos, NTLM, and lateral movement testing.', 'smbexec.py; secretsdump.py; wmiexec.py'],
        ['BURP SUITE', 'Web application interception, fuzzing, and vulnerability validation platform.', 'proxy; repeater; intruder'],
        ['', '', ''],
        ['--- RED TEAM STRATEGY: ADVERSARIAL EMULATION ---', '', ''],
        ['FRAMEWORK', 'FOCUS', 'CORE METRIC'],
        ['MITRE ATT&CK', 'Mapping techniques to the Matrix.', 'T1059.001 (PowerShell)'],
        ['PTES', 'Standardized pentest execution.', 'Exploitation Success Rate'],
        ['OSSTMM', 'Operational security methodology.', 'RAV (Risk AssessmentValue)'],
        ['OWASP TOP 10', 'Web application risk categories and validation targets.', 'A01, A03, A05 coverage'],
        ['CYBER KILL CHAIN', 'Intrusion lifecycle model from recon through actions on objectives.', 'Phase completion and detection point'],
        ['', '', ''],
        ['--- DETECTION MATURITY LEVELS ---', '', ''],
        ['LEVEL', 'DESCRIPTION', 'CAPABILITY'],
        ['1 - INITIAL', 'Logs exist but are not centralized.', 'Manual grepping of logs.'],
        ['2 - MANAGED', 'Core telemetry is collected from priority systems.', 'Repeatable alert review.'],
        ['3 - DEFINED', 'Centralized alerts for known IOCs.', 'Security Onion/Suricata alerting.'],
        ['4 - MEASURED', 'Detection latency, coverage, and false positive rates are tracked.', 'Dashboards and hunt metrics.'],
        ['5 - OPTIMIZED', 'Automated response and behavioral hunting.', 'Custom Playbooks/SOAR.'],
        ['', '', ''],
        ['--- PURPLE TEAM ENGAGEMENT WORKFLOW ---', '', ''],
        ['STAGE', 'JOINT ACTIVITY', 'DELIVERABLE'],
        ['PLAN', 'Define scope, rules of engagement, target systems, and success criteria.', 'Engagement plan and test matrix.'],
        ['EMULATE', 'Red team executes selected ATT&CK techniques with controlled timing.', 'Command log and attack timeline.'],
        ['DETECT', 'Blue team validates alerts, telemetry, and investigation workflow.', 'Detection notes and evidence IDs.'],
        ['TUNE', 'Refine SIEM rules, Suricata signatures, Sysmon config, and response playbooks.', 'Updated detections and mitigations.'],
        ['VALIDATE', 'Re-run tests and compare detection latency, containment time, and coverage.', 'Before/after validation results.'],
        ['', '', ''],
        ['--- MITRE ATT&CK COVERAGE PLAN ---', '', ''],
        ['TACTIC', 'TECHNIQUE FOCUS', 'EXPECTED BLUE TEAM SIGNAL'],
        ['Initial Access', 'T1190 Exploit Public-Facing Application', 'Web logs, Suricata alerts, anomalous process launch.'],
        ['Execution', 'T1059 Command and Scripting Interpreter', 'PowerShell/script telemetry and command-line logging.'],
        ['Lateral Movement', 'T1021.001 SMB/Windows Admin Shares', 'SMB session telemetry, Windows Event IDs, Zeek notices.'],
        ['Credential Access', 'T1003 OS Credential Dumping', 'LSASS access, suspicious handle events, EDR alerts.'],
        ['Exfiltration', 'T1041 Exfiltration Over C2 Channel', 'Outbound volume anomaly, protocol mismatch, destination reputation.'],
        ['', '', ''],
        ['--- RULES OF ENGAGEMENT ---', '', ''],
        ['CONTROL', 'REQUIREMENT', 'OWNER'],
        ['Authorization', 'Operate only inside approved lab networks and authorized assets.', 'Engagement Lead'],
        ['Safety', 'Avoid destructive payloads, persistence outside the lab, and uncontrolled propagation.', 'Red Team Lead'],
        ['Evidence', 'Record commands, timestamps, hashes, screenshots, and packet captures.', 'Both Teams'],
        ['Stop Conditions', 'Pause testing if production systems, personal data, or unstable hosts are affected.', 'Engagement Lead'],
        ['', '', ''],
        ['--- SUCCESS METRICS ---', '', ''],
        ['METRIC', 'TARGET', 'MEASUREMENT SOURCE'],
        ['Mean Time to Detect', '< 30 minutes', 'Blue_Team_Logs and Visual_Analytics'],
        ['Mean Time to Contain', '< 60 minutes', 'Incident timeline and containment actions'],
        ['Detection Coverage', '>= 80% of tested techniques', 'MITRE coverage matrix'],
        ['Evidence Integrity', '100% hash verification for retained artifacts', 'Evidence_Vault'],
        ['Remediation Closure', 'All critical/high findings assigned and tracked', 'Lessons_Learned and OWASP tabs'],
    ]

    # TAB 7: BLUE TEAM (Persona: Ebony Rose)
    blue_data = {
        'Event_Date': [today_str],
        'Event_Time': [now_ts],
        'Discovery_Time': [''],
        'Analyst': [''],
        'NIST_Lifecycle_Phase': ['Detection & Analysis'],
        'Incident_Category': ['Unauthorized Access'],
        'Red_Event_Link': ['RT-001'],  # MISSING (The "Purple" Link)
        'CSF_Function': ['Respond'], 'NIST_Phase': ['Detection & Analysis'],
        'Disposition': ['True Positive'],
        'Detection_Source': ['Security Onion (Suricata)'],
        'Source_MAC': [''],
        'Source_IP': ['192.168.30.10'],  # MISSING
        'Target_IP/VLAN': ['192.168.x.x'],
        'Target_Host': ['192.168.20.5'],
        'Process_Name': ['java.exe'],  # MISSING
        'CVE_Ref (NEW)': ['CVE-2021-44228'],
        'Detection_Latency': ['45s'],  # MISSING
        'Impact_Level': ['Medium'],
        'Containment_Action': ['VLAN Isolation'],
        'Containment_Time': ['15:04:22'],
        'Risk_Impact': ['Critical'],
        'Firewall_Action (NEW)': ['Blocked'],
        'Command_Executed': ['so-status'],
        'Status': ['Closed'],
        'Resolution_Status': ['Active'],
        'Evidence_Ref': ['EV-001'],
        'Findings/Observations': ['Logged'],
        'Analysis_Notes': ['Confirmed JNDI string in PCAP']  # MISSING
    }

    # TAB 8: RED TEAM (Restored Multi-Event Log)
    red_events = [
        {
            'Date': today_str,
            'Timestamp': now_ts,
            'Analyst': '',
            'Campaign_Objective': 'Initial Access',
            'PTES_Phase': 'Exploitation',
            'OSSTMM_Class': 'Data Networks',
            'OPSEC_Risk': 'High',
            'MITRE_ID': 'T1190',
            'MITRE_ATT&CK_ID': 'T1190',
            'Tool_Used': 'Metasploit',
            'Command_Executed': 'exploit/multi/http/log4shell',
            'Methodology': 'Service Mapping',
            'Raw_Tool_Output': 'Exploit attempt logged',
            'Operator_Visibility': 'Noisy',
            'Target_Host': '192.168.20.5',
            'Privilege_Escalation': 'N/A',
            'C2_Infrastructure': 'Cobalt Strike',
            'Egress_Protocol': 'HTTP',
            'Risk_Score': 'High',
            'Success': 'True'
        },
        {
            'Date': today_str,
            'Timestamp': now_ts,
            'Analyst': '',
            'Campaign_Objective': 'Lateral Movement',
            'PTES_Phase': 'Post-Exploitation',
            'OSSTMM_Class': 'Data Networks',
            'OPSEC_Risk': 'Medium',
            'MITRE_ID': 'T1021.001',
            'MITRE_ATT&CK_ID': 'T1021.001',
            'Tool_Used': 'Impacket / SMBExec',
            'Command_Executed': 'python3 smbexec.py admin:pass@10.0.0.1',
            'Methodology': 'Remote service execution',
            'Raw_Tool_Output': 'Administrative session established',
            'Operator_Visibility': 'Stealthy',
            'Target_Host': '192.168.20.5',
            'Privilege_Escalation': 'Local Admin',
            'C2_Infrastructure': 'Cobalt Strike',
            'Egress_Protocol': 'SMB over VPN',
            'Risk_Score': 'Medium',
            'Success': 'True'
        },
        {
            'Date': today_str,
            'Timestamp': now_ts,
            'Analyst': '',
            'Campaign_Objective': 'Service Mapping',
            'PTES_Phase': 'Exploitation',
            'OSSTMM_Class': 'Data Networks',
            'OPSEC_Risk': 'Medium',
            'MITRE_ID': 'T1059.001',
            'MITRE_ATT&CK_ID': 'T1059.001',
            'Tool_Used': 'Nmap',
            'Command_Executed': 'nmap -sV --script vuln',
            'Methodology': 'Service Mapping',
            'Raw_Tool_Output': 'Filtered',
            'Operator_Visibility': 'Noisy',
            'Target_Host': '192.168.x.x',
            'Privilege_Escalation': 'N/A',
            'C2_Infrastructure': 'N/A',
            'Egress_Protocol': 'N/A',
            'Risk_Score': 'Medium',
            'Success': 'True'
        }
    ]

    # TAB 9: OWASP (With Red Banner Mapping)
    owasp_data = {
        'Date': [today_str, today_str],
        'Timestamp': [now_ts, now_ts],
        'Analyst': ['', ''],
        'Vuln_ID': ['A01:2021', 'A03:2021'],
        'OWASP_Category': ['A01:2021-Broken Access Control', 'A03:2021-Injection'],
        'Category': ['Broken Access Control', 'Injection'],
        'CWE_Mapping': ['CWE-639 (IDOR)', 'CWE-89 (SQL Injection)'],
        'Target_URL': ['192.168.x.x/admin', 'Login Field'],
        'Description': [
            'Insecure Direct Object Reference (IDOR) flaw in /user/profile endpoints.',
            'Input handling permits injection testing against login parameters.'
        ],
        'Severity': ['High', 'High'],
        'Likelihood': ['Medium', 'Medium'],
        'Business_Impact': ['Unauthorized access to PII and potential full account takeover.', 'Unauthorized data access.'],
        'Technical_Impact': ['Vertical Privilege Escalation', 'Database query manipulation'],
        'Command_Executed': ['ffuf', 'sqlmap'],
        'Findings/Observations': ['Critical', 'High'],
        'PoC_Reference': ['Artifact EV-002 / Request: GET /api/v1/user/profile?id=1002', 'Artifact EV-003 / Login test'],
        'Remediation': ['Implement JWT-based claim validation and server-side ownership checks.', 'Parameterized Queries'],
        'Remediation_Plan': ['Implement RBAC', 'Parameterized Queries'],
        'Owner': ['Web_Development_Team', 'Web_Development_Team'],
        'Status': ['Open', 'Open'],
        'Discovery_Date': [today_str, today_str]
    }

    owasp_df = pd.DataFrame(owasp_data)

    # TAB 10: LESSONS LEARNED
    lessons_learned_rows = [
        ['Issue Identified', 'Root Cause', 'Detection Gap', 'MITRE Ref', 'Remediation Action', 'Complexity',
         'Risk Reduction', 'Validation Test', 'Owner', 'Status'],
        ['Log4j alerts delayed', 'ES indexing lag', 'Process', 'T1190', 'Optimize Indexing', 'Medium', 'High',
         'Re-run Log4Shell exploit', 'DevOps', 'Pending'],
        ['SMBExec undetected', 'Missing Sysmon EID 1', 'Visibility', 'T1021.001', 'Update Sysmon Config', 'Low',
         'Critical', 'Invoke-AtomicTest T1021.001', 'Blue Team', 'Open']
    ]

    # TAB 11: METRICS
    det_metrics = pd.DataFrame({'Category': ['Detected', 'Missed', 'Blocked'], 'Count': [15, 3, 7]})
    maturity_metrics = pd.DataFrame(
        {'Metric': ['Detection', 'Response', 'Logs', 'Contain', 'Integrate'], 'Score': [8, 6, 9, 7, 8]})
    trend_metrics = pd.DataFrame(
        {'Session': ['Wk 1', 'Wk 2', 'Wk 3', 'Wk 4'], 'MTTD (Min)': [45, 30, 15, 12], 'MTTC (Min)': [120, 90, 45, 30]})


    # --- 3. CHARTING SUB-FUNCTION ---
    # DETECTION METRICS (Original description)
    det_metrics = pd.DataFrame({
        'Category': ['Detected', 'Missed', 'Blocked'],
        'Count': [15, 3, 7]
    })

    # MATURITY METRICS (Original description)
    maturity_metrics = pd.DataFrame({
        'Metric': ['Detection', 'Response', 'Logs', 'Contain', 'Integrate'],
        'Score': [8, 6, 9, 7, 8]
    })

    # EFFICACY METRICS (Original description)
    efficacy_metrics = pd.DataFrame({
        'Phase': ['Prep', 'Detect', 'Contain', 'Eradicate'],
        'Efficacy': [90, 75, 85, 95]
    })

    # MITRE COVERAGE (Strategic Addition)
    mitre_coverage = pd.DataFrame({
        'Tactic': ['Initial Access', 'Execution', 'Persistence', 'Exfiltration'],
        'Techniques_Tested': [5, 4, 2, 3],
        'Coverage_Score (%)': [100, 75, 20, 50]
    })

    # TIME TRENDS (Performance tracking for line charts)
    trend_metrics = pd.DataFrame({
        'Session': ['Wk 1', 'Wk 2', 'Wk 3', 'Wk 4'],
        'MTTD (Min)': [45, 30, 15, 12],
        'MTTC (Min)': [120, 90, 45, 30]
    })

    # --- 3. FIXED CHARTING FUNCTION ---
    def write_analytics_tab(writer, det, mat, eff, mitre, trend):
        """Generates the Visual Analytics tab using passed metrics."""

        det.to_excel(writer, sheet_name='Visual_Analytics', index=False, startrow=2, startcol=1)
        mat.to_excel(writer, sheet_name='Visual_Analytics', index=False, startrow=2, startcol=5)
        eff.to_excel(writer, sheet_name='Visual_Analytics', index=False, startrow=2, startcol=9)
        mitre.to_excel(writer, sheet_name='Visual_Analytics', index=False, startrow=2, startcol=13)
        trend.to_excel(writer, sheet_name='Visual_Analytics', index=False, startrow=2, startcol=17)

        ws = writer.sheets['Visual_Analytics']
        headers = [("B2", "DETECTION"), ("F2", "MATURITY"), ("J2", "EFFICACY"), ("N2", "MITRE"), ("R2", "TRENDS")]
        for cell, val in headers:
            ws[cell] = val
            ws[cell].fill = PatternFill(start_color=PURPLE_THEME, end_color=PURPLE_THEME, fill_type="solid")
            ws[cell].font = Font(color="FFFFFF", bold=True)
            ws[cell].alignment = Alignment(horizontal="center")

        # Add Charts
        doughnut = DoughnutChart()
        doughnut.title = "Detection Breakdown"
        doughnut.holeSize = 55
        doughnut.firstSliceAng = 270
        doughnut.add_data(Reference(ws, min_col=3, min_row=3, max_row=6), titles_from_data=True)
        doughnut.set_categories(Reference(ws, min_col=2, min_row=4, max_row=6))
        doughnut.width = 12
        doughnut.height = 8
        ws.add_chart(doughnut, "B11")

        radar = RadarChart();
        radar.title = "NIST Maturity Radar"
        radar.style = 26
        radar.add_data(Reference(ws, min_col=7, min_row=3, max_row=8), titles_from_data=True)
        radar.set_categories(Reference(ws, min_col=6, min_row=4, max_row=8))
        radar.width = 12
        radar.height = 8
        ws.add_chart(radar, "J11")

        bar = BarChart();
        bar.title = "ATT&CK Coverage"
        bar.style = 10
        bar.add_data(Reference(ws, min_col=15, min_row=3, max_row=7), titles_from_data=True)
        bar.set_categories(Reference(ws, min_col=14, min_row=4, max_row=7))
        bar.width = 14
        bar.height = 8
        ws.add_chart(bar, "B28")

        eff_bar = BarChart()
        eff_bar.title = "Response Efficacy"
        eff_bar.type = "bar"
        eff_bar.style = 12
        eff_bar.add_data(Reference(ws, min_col=11, min_row=3, max_row=7), titles_from_data=True)
        eff_bar.set_categories(Reference(ws, min_col=10, min_row=4, max_row=7))
        eff_bar.width = 12
        eff_bar.height = 8
        ws.add_chart(eff_bar, "J28")

        area = AreaChart()
        area.title = "Response Latency Trend"
        area.style = 13
        area.add_data(Reference(ws, min_col=18, min_row=3, max_col=19, max_row=7), titles_from_data=True)
        area.set_categories(Reference(ws, min_col=17, min_row=4, max_row=7))
        area.width = 14
        area.height = 8
        ws.add_chart(area, "B45")
        return

    def apply_sheet_banner(ws, title, color, banner_rows=1):
        """Apply a full-width banner and readable column headers."""
        max_col = max(ws.max_column, 1)
        ws.merge_cells(start_row=1, start_column=1, end_row=banner_rows, end_column=max_col)
        cell = ws.cell(row=1, column=1)
        cell.value = title
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.font = Font(color="000000" if color == LILAC_PURPLE else "FFFFFF", bold=True, size=16 if banner_rows > 1 else 14)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for row_idx in range(1, banner_rows + 1):
            ws.row_dimensions[row_idx].height = 24

        header_row = banner_rows + 1
        for header in ws[header_row]:
            if not isinstance(header, MergedCell) and header.value:
                header_color = "D9EAF7"
                if ws.title in ("Blue_Team_Logs", "Defensive_Logs"):
                    header_color = BLUE_TEAM_LIGHT
                elif ws.title in ("Red_Team_Logs", "Offensive_Logs", "OWASP"):
                    header_color = RED_TEAM_LIGHT
                elif ws.title == "Purple_Strategy":
                    header_color = LILAC_PURPLE
                header.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
                header.font = Font(bold=True)
                header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def strategy_section_style(section_title):
        text = str(section_title).upper()
        if "BLUE TEAM ENCYCLOPEDIA" in text:
            return BLUE_TEAM_COLOR, "FFFFFF"
        if "BLUE TEAM" in text:
            return BLUE_TEAM_LIGHT, "000000"
        if "RED TEAM ENCYCLOPEDIA" in text:
            return RED_TEAM_COLOR, "FFFFFF"
        if "RED TEAM" in text:
            return RED_TEAM_LIGHT, "000000"
        if "PURPLE" in text or "MITRE" in text or "DETECTION MATURITY" in text:
            return PURPLE_THEME, "FFFFFF"
        return LILAC_PURPLE, "000000"

    def apply_section_banners(ws, start_row=2):
        """Style in-sheet section labels such as --- CASE IDENTIFICATION ---."""
        max_col = 3 if ws.title == "Purple_Strategy" else max(ws.max_column, 1)
        for row in range(start_row, ws.max_row + 1):
            cell = ws.cell(row=row, column=1)
            if cell.value and "---" in str(cell.value):
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
                fill_color, font_color = (
                    strategy_section_style(cell.value)
                    if ws.title == "Purple_Strategy"
                    else (LILAC_PURPLE, "000000")
                )
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.font = Font(color=font_color, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

    def add_purple_strategy_legend(ws):
        """Place an operational legend to the right of the Purple Strategy content."""
        legend_rows = [
            ("Section", "Use this for", "Owner", "Record output in", PURPLE_THEME, "FFFFFF"),
            ("Blue Team Encyclopedia", "Look up defensive tools, expected telemetry, and command references.", "Blue Team", "Blue_Team_Logs / Defensive_Logs", BLUE_TEAM_COLOR, "FFFFFF"),
            ("Blue Team Strategy", "Map observations to NIST incident response phases and response goals.", "Blue Team", "Dossier / Lessons_Learned", BLUE_TEAM_MID, "FFFFFF"),
            ("Red Team Encyclopedia", "Look up offensive tools, commands, and emulation techniques.", "Red Team", "Red_Team_Logs / Offensive_Logs", RED_TEAM_COLOR, "FFFFFF"),
            ("Red Team Strategy", "Map actions to ATT&CK, PTES, OSSTMM, and OWASP coverage.", "Red Team", "Red_Team_Logs / OWASP", RED_TEAM_MID, "FFFFFF"),
            ("Detection Maturity", "Rate whether telemetry is manual, centralized, measured, or automated.", "Purple Team", "Visual_Analytics / Lessons_Learned", PURPLE_THEME, "FFFFFF"),
            ("Engagement Workflow", "Follow the operating sequence: plan, emulate, detect, tune, validate.", "Both Teams", "Dossier / Evidence_Vault", PURPLE_THEME, "FFFFFF"),
            ("MITRE Coverage Plan", "Choose techniques to test and define the expected defensive signal.", "Both Teams", "Visual_Analytics / Blue_Team_Logs", PURPLE_THEME, "FFFFFF"),
            ("Rules of Engagement", "Check authorization, safety limits, evidence rules, and stop conditions.", "Engagement Lead", "Dossier", DARK_GREY, "FFFFFF"),
            ("Success Metrics", "Track detection speed, containment speed, coverage, evidence integrity, and closure.", "Engagement Lead", "Dossier / Visual_Analytics", NAVY_BLUE, "FFFFFF"),
        ]
        start_row, start_col = 3, 8
        end_col = start_col + 3
        ws.column_dimensions[get_column_letter(start_col)].width = 24
        ws.column_dimensions[get_column_letter(start_col + 1)].width = 44
        ws.column_dimensions[get_column_letter(start_col + 2)].width = 18
        ws.column_dimensions[get_column_letter(start_col + 3)].width = 30
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
        title_cell = ws.cell(row=start_row, column=start_col)
        title_cell.value = "LEGEND"
        title_cell.fill = PatternFill(start_color=PURPLE_THEME, end_color=PURPLE_THEME, fill_type="solid")
        title_cell.font = Font(color="FFFFFF", bold=True, size=13)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        for offset, (section, purpose, owner, output, fill_color, font_color) in enumerate(legend_rows):
            row = start_row + offset + 1
            values = [section, purpose, owner, output]
            for col_offset, value in enumerate(values):
                cell = ws.cell(row=row, column=start_col + col_offset)
                cell.value = value
                cell.alignment = Alignment(
                    horizontal="center" if offset == 0 or col_offset in (0, 2) else "left",
                    vertical="center",
                    wrap_text=True
                )
                if offset == 0:
                    cell.fill = PatternFill(start_color=PURPLE_THEME, end_color=PURPLE_THEME, fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                elif col_offset == 0:
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                else:
                    cell.font = Font(bold=False, color="000000")

        legend_border = Border(
            left=Side(style="medium", color="000000"),
            right=Side(style="medium", color="000000"),
            top=Side(style="medium", color="000000"),
            bottom=Side(style="medium", color="000000"),
        )
        inner_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )
        end_row = start_row + len(legend_rows)
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = inner_border
        for col in range(start_col, end_col + 1):
            ws.cell(row=start_row, column=col).border = legend_border
            ws.cell(row=end_row, column=col).border = legend_border
        for row in range(start_row, end_row + 1):
            ws.cell(row=row, column=start_col).border = legend_border
            ws.cell(row=row, column=end_col).border = legend_border

    def normalize_purple_strategy_data_rows(ws):
        """Keep non-header strategy rows plain and left-aligned."""
        header_labels = {"TOOL", "PHASE", "FRAMEWORK", "LEVEL", "STAGE", "TACTIC", "CONTROL", "METRIC"}
        for row_idx in range(3, ws.max_row + 1):
            first_value = ws.cell(row=row_idx, column=1).value
            first_text = str(first_value).upper() if first_value is not None else ""
            if not first_value or "---" in first_text or first_text in header_labels:
                continue
            for col_idx in range(1, 4):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell, MergedCell):
                    continue
                cell.fill = PatternFill(fill_type=None)
                cell.font = Font(bold=False, color="000000")
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    def improve_readability(ws):
        """Set practical widths and wrapping so long workbook text remains visible."""
        category_labels = {
            "TOOL", "PHASE", "FRAMEWORK", "LEVEL", "STAGE", "TACTIC", "CONTROL", "METRIC"
        }
        header_rows = {2}
        if ws.title == "Purple_Strategy":
            header_rows = {
                row_idx
                for row_idx in range(1, ws.max_row + 1)
                if str(ws.cell(row=row_idx, column=1).value).upper() in category_labels
            }
        elif ws.title in ("Lessons_Learned",):
            header_rows = {2}

        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                is_banner_row = cell.row == 1 or (ws.title == "Purple_Strategy" and cell.row == 2)
                is_category_row = cell.row in header_rows and cell.value is not None
                cell.alignment = Alignment(
                    horizontal="center" if is_banner_row or is_category_row else (cell.alignment.horizontal or "left"),
                    vertical="center" if is_banner_row else "top",
                    wrap_text=True
                )
                if is_category_row:
                    cell.font = Font(bold=True)

        for column_cells in ws.columns:
            col_letter = get_column_letter(column_cells[0].column)
            max_length = 0
            for cell in column_cells:
                if isinstance(cell, MergedCell) or cell.value is None:
                    continue
                max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(max_length + 3, 14), 48)

        for row_idx in range(1, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 30 if row_idx == 1 else 42

    def apply_black_gridlines(ws):
        """Apply visible black cell borders across each populated worksheet range."""
        border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border

    def resolve_output_filename(path):
        """Use the requested path unless Excel has it locked, then write a timestamped copy."""
        if not os.path.exists(path):
            return path
        try:
            with open(path, "a+b"):
                return path
        except PermissionError:
            root, ext = os.path.splitext(path)
            fallback = f"{root}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
            print(f"Workbook is open or locked. Creating a new copy: {fallback}")
            return fallback

    # --- 4. THE WRITER & FORMATTING BLOCK ---
    try:
        filename = resolve_output_filename(filename)
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # --- Tab 1: Dossier ---
            pd.DataFrame(dossier_rows, columns=['Field', 'Value', 'Threshold']).to_excel(
                writer, sheet_name='Dossier', index=False, startrow=1
            )

            # --- Tab 2: Evidence Vault ---
            pd.DataFrame(ev_data).to_excel(writer, sheet_name='Evidence_Vault', index=False, startrow=1)

            # --- Tab 3: Forensic Evaluation ---
            pd.DataFrame(fe_data).to_excel(writer, sheet_name='Forensic_Evaluation', index=False, startrow=1)

            # --- Tab 4: Topology ---
            pd.DataFrame(topo_rows, columns=topo_cols).to_excel(
                writer, sheet_name='Topology', index=False, startrow=1
            )

            # --- Tab 5: Purple Strategy ---
            pd.DataFrame(ps_rows).to_excel(
                writer, sheet_name='Purple_Strategy', index=False, header=False, startrow=2
            )

            # --- Tab 6: Blue Team ---
            pd.DataFrame(blue_data).to_excel(writer, sheet_name='Blue_Team_Logs', index=False, startrow=1)

            # --- Tab 7: Red Team ---
            pd.DataFrame(red_events).to_excel(writer, sheet_name='Red_Team_Logs', index=False, startrow=1)

            # --- Tab 8: OWASP ---
            owasp_df.to_excel(writer, sheet_name='OWASP', index=False, startrow=1)

            # --- Tab 9: Lessons Learned ---
            pd.DataFrame(lessons_learned_rows).to_excel(
                writer, sheet_name='Lessons_Learned', index=False, header=False, startrow=1
            )

            # --- Tab 10: Metrics / Analytics ---
            write_analytics_tab(writer, det_metrics, maturity_metrics, efficacy_metrics, mitre_coverage, trend_metrics)

            # --- Additional Logging Templates ---
            def_cols = ['Date', 'Timestamp', 'Analyst', 'Phase', 'Category', 'Source MAC', 'Command', 'Findings']
            pd.DataFrame(columns=def_cols).to_excel(writer, sheet_name='Defensive_Logs', index=False, startrow=1)

            off_cols = ['Date', 'Timestamp', 'Analyst', 'Phase', 'MITRE ID', 'Target Host', 'Command', 'Risk Score']
            pd.DataFrame(columns=off_cols).to_excel(writer, sheet_name='Offensive_Logs', index=False, startrow=1)

            # --- 5. BANNER FORMATTING ---
            config = [
                ('Dossier', "SYSTEM DOSSIER", DARK_GREY),
                ('Evidence_Vault', "EVIDENCE CHAIN OF CUSTODY", NAVY_BLUE),
                ('Forensic_Evaluation', "FORENSIC ANALYSIS LOG", DARK_GREY),
                ('Topology', "NETWORK INFRASTRUCTURE", NAVY_BLUE),
                ('Purple_Strategy', "STRATEGY & TOOL ENCYCLOPEDIA", PURPLE_THEME),
                ('Blue_Team_Logs', "BLUE TEAM ACTIVITY", BLUE_TEAM_MID),
                ('Red_Team_Logs', "RED TEAM ACTIVITY", RED_TEAM_MID),
                ('OWASP', "VULNERABILITY ASSESSMENT", RED_TEAM_MID),
                ('Lessons_Learned', "POST-ENGAGEMENT LESSONS", NAVY_BLUE),
                ('Visual_Analytics', "EXECUTIVE ANALYTICS DASHBOARD", DARK_GREY),
                ('Defensive_Logs', "ACTIVE DEFENSE MONITORING", BLUE_TEAM_MID),
                ('Offensive_Logs', "ADVERSARIAL EMULATION LOGS", RED_TEAM_MID)
            ]

            for sheet_name, title, color in config:
                if sheet_name in writer.sheets:
                    ws = writer.sheets[sheet_name]
                    banner_rows = 2 if sheet_name == "Purple_Strategy" else 1
                    apply_sheet_banner(ws, title, color, banner_rows=banner_rows)
                    apply_section_banners(ws, start_row=banner_rows + 1)
                    improve_readability(ws)
                    if sheet_name == "Purple_Strategy":
                        normalize_purple_strategy_data_rows(ws)
                    apply_black_gridlines(ws)
                    if sheet_name == "Purple_Strategy":
                        add_purple_strategy_legend(ws)
                    ws.sheet_properties.tabColor = color

            for sheet in writer.sheets.values():
                if sheet.title not in {sheet_name for sheet_name, _, _ in config}:
                    improve_readability(sheet)
                    apply_black_gridlines(sheet)

        print(f"Success: {filename} has been created.")
        if os.path.exists(filename):
            os.startfile(filename)
            print("Opening Excel...")
        return filename

    except PermissionError:
        print(f"CRITICAL ERROR: Please close '{filename}' before running.")
        raise

    except Exception as e:
        print(f"Error during workbook generation: {e}. If the workbook is open in Excel, close it and run again.")
        raise


if __name__ == "__main__":
    purple_team_workbook()
