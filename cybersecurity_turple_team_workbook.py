import pandas as pd
import os
from datetime import datetime
from openpyxl.styles import Alignment, Font, PatternFill


def generate_full_framework_master_workbook():
    # 1. ROBUST PATH DISCOVERY
    user_profile = os.environ['USERPROFILE']
    paths_to_check = [
        os.path.join(user_profile, 'OneDrive', 'Desktop'),
        os.path.join(user_profile, 'Desktop'),
        user_profile
    ]

    desktop_path = next((p for p in paths_to_check if os.path.exists(p)), user_profile)
    filename = os.path.join(desktop_path, "Cybersecurity_Unified_Operations_Workbook.xlsx")

    today_str = datetime.now().strftime('%Y-%m-%d')
    now_ts = datetime.now().strftime('%H:%M:%S')

    # --- 2. DATA STRUCTURES ---

    # ADMINISTRATIVE: Dossier (Length 6)
    dossier_data = {
        'Field': ['Case ID', 'Analyst', 'Sector', 'Classification', 'Standards', 'Lab Environment'],
        'Value': [f'CASE-{datetime.now().year}-001', 'First Last Name', 'Cybersecurity Research', 'Confidential',
                  'NIST / PTES / OSSTMM / MITRE', 'Security Onion / pfSense / Netgate']
    }

    # 🛡 BLUE TEAM: NIST 800-61 (Length 1)
    blue_team_data = {
        'Event_Date': [today_str],
        'Event_Time': [now_ts],
        'Discovery_Time': [''],
        'Analyst': ['First Last Name'],
        'NIST_Lifecycle_Phase': ['Detection & Analysis'],
        'Incident_Category': ['Unauthorized Access'],
        'Detection_Source': ['Security Onion (Zeek/Suricata)'],
        'Source_MAC': [''],
        'Target_IP/VLAN': ['192.168.x.x'],
        'Impact_Level': ['Medium'],
        'Containment_Action': ['VLAN Isolation'],
        'Command_Executed': ['so-status'],
        'Findings/Observations': ['Logged'],
        'Resolution_Status': ['Active']
    }

    # ⚔️ RED TEAM: PTES & OSSTMM (Length 1)
    red_team_data = {
        'Date': [today_str],
        'Timestamp': [now_ts],
        'Analyst': ['First Last Name'],
        'PTES_Phase': ['Exploitation'],
        'OSSTMM_Class': ['Data Networks'],
        'MITRE_ATT&CK_ID': ['T1059.001'],
        'Target_Host': ['192.168.x.x'],
        'Methodology': ['Service Mapping'],
        'Command_Executed': ['nmap -sV --script vuln'],
        'Raw_Tool_Output': ['Filtered'],
        'Risk_Score': ['Medium']
    }

    # RED TEAM: Web App OWASP (Length 2)
    owasp_data = {
        'Date': [today_str, today_str],
        'Timestamp': [now_ts, now_ts],
        'Analyst': ['First Last Name', 'First Last Name'],  # FIXED: Balanced to length 2
        'OWASP_Category': ['A01:2021-Broken Access Control', 'A03:2021-Injection'],
        'Target_URL': ['192.168.x.x/admin', 'Login Field'],
        'Command_Executed': ['ffuf', 'sqlmap'],
        'Findings/Observations': ['Critical', 'High'],
        'Remediation_Plan': ['Implement RBAC', 'Parameterized Queries']
    }

    # ⛓️ EVIDENCE VAULT (Length 2)
    evidence_vault = {
        'Date': [today_str, today_str],
        'Timestamp': [now_ts, now_ts],
        'Evidence_ID': ['EV-001', 'EV-002'],
        'Artifact_Source': ['Nmap Scan Log', 'Wireshark PCAP'],
        'SHA-256_Hash': ['e3b0c442...', '5df6e0e2...'],
        'Analyst': ['First Last Name', 'First Last Name'],  # FIXED: Balanced to length 2
        'Integrity_Status': ['Verified', 'Verified']
    }

    # --- 3. GENERATION & STYLING ---
    try:
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            pd.DataFrame(dossier_data).to_excel(writer, sheet_name='Dossier', index=False, startrow=1)
            pd.DataFrame(blue_team_data).to_excel(writer, sheet_name='Blue_Team_NIST_IR', index=False, startrow=1)
            pd.DataFrame(red_team_data).to_excel(writer, sheet_name='Red_Team_PTES_OSSTMM', index=False, startrow=1)
            pd.DataFrame(owasp_data).to_excel(writer, sheet_name='Red_Team_OWASP', index=False, startrow=1)
            pd.DataFrame(evidence_vault).to_excel(writer, sheet_name='Evidence_Vault', index=False, startrow=1)

            # Define styling parameters for each sheet
            for sheet_name, title, color, col_count in [
                ('Dossier', "INVESTIGATION DOSSIER & FRAMEWORK MAPPING", "212121", 2),
                ('Blue_Team_NIST_IR', "DEFENSIVE OPERATIONS: NIST IR & SIEM CORRELATION", "0D47A1", 14),
                ('Red_Team_PTES_OSSTMM', "OFFENSIVE OPERATIONS: MITRE / PTES / OSSTMM", "B71C1C", 11),
                ('Red_Team_OWASP', "RED TEAM: OWASP TOP 10 WEB ASSESSMENT", "B71C1C", 8),
                ('Evidence_Vault', "FORENSIC CHAIN OF CUSTODY (SHA-256)", "424242", 7)
            ]:
                ws = writer.sheets[sheet_name]
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
                header = ws.cell(row=1, column=1)
                header.value = title
                header.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                header.font = Font(color="FFFFFF", bold=True, size=12)
                header.alignment = Alignment(horizontal="center")

                for col in range(1, col_count + 1):
                    col_header = ws.cell(row=2, column=col)
                    col_header.font = Font(bold=True)
                    col_header.alignment = Alignment(horizontal="center")

        print(f"SUCCESS: Workbook generated at {filename}")
        os.startfile(filename)

    except Exception as e:
        print(f"SYSTEM ERROR: {e}")


if __name__ == "__main__":
    generate_full_framework_master_workbook()
