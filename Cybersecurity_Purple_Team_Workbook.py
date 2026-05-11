import pandas as pd
import os
from datetime import datetime
from openpyxl.styles import Alignment, Font, PatternFill


def generate_ultimate_master_workbook():
    # 1. ROBUST PATH DISCOVERY (Fixes the 'non-existent directory' error)
    user_profile = os.environ['USERPROFILE']
    paths_to_check = [
        os.path.join(user_profile, 'OneDrive', 'Desktop'),
        os.path.join(user_profile, 'Desktop'),
        user_profile
    ]

    desktop_path = next((p for p in paths_to_check if os.path.exists(p)), user_profile)
    filename = os.path.join(desktop_path, "Cybersecurity_Purple_Team_Investigation_Workbook.xlsx")

    today_str = datetime.now().strftime('%Y-%m-%d')
    now_ts = datetime.now().strftime('%H:%M:%S')

    # --- 2. DATA STRUCTURES (Re-Integrated Features) ---

    # NEW: Dossier Tab (Restored from Version 1)
    dossier_data = {
        'Field': ['Case ID', 'Analyst', 'Sector', 'Classification', 'Standard', 'Lab Environment'],
        'Value': ['PHX-2026-001', 'First Last Name', 'Cybersecurity Research', 'Confidential', 'NIST / PTES / OSSTMM',
                  'Phoenix Prime Lab']
    }

    # BLUE TEAM: NIST Incident Response (Tactical + NIST Phases restored)
    blue_team_data = {
        'Date': [today_str],
        'Timestamp': [now_ts],
        'Analyst': ['First Last Name'],
        'NIST_Phase': ['Detection/Analysis'],  # Restored explicit phase
        'Incident_Type': ['Unauthorized Access'],
        'Detection_Source': ['Snort/Security Onion'],
        'Source_MAC': [''],
        'Target_IP': ['192.168.x.x'],
        'TTL_Value': [''],
        'Impact_Level': ['Medium'],
        'Command_Executed': [''],
        'Findings/Observations': ['Logged'],
        'Resolution_Status': ['Active']
    }

    # RED TEAM: PTES & OSSTMM (Restored OSSTMM Class & Methodology)
    red_team_data = {
        'Date': [today_str],
        'Timestamp': [now_ts],
        'Analyst': ['First Last Name'],
        'PTES_Phase': ['Intelligence Gathering'],
        'OSSTMM_Class': ['Data Networks'],  # Restored OSSTMM Class
        'Methodology': ['Service Mapping'],  # Restored Methodology
        'Target_Host': ['192.168.x.x'],
        'Command_Executed': ['nmap -sV'],
        'Findings/Observations': ['Filtered'],
        'Risk_Score': ['Medium']
    }

    # RED TEAM: OWASP
    owasp_data = {
        'Date': [today_str, today_str],
        'Timestamp': [now_ts, now_ts],
        'OWASP_Category': ['A01:2021-Broken Access Control', 'A03:2021-Injection'],
        'Target_URL': ['192.168.x.x/admin', 'Login Field'],
        'Command_Executed': ['', ''],
        'Findings/Observations': ['Critical', 'High'],
        'Remediation_Plan': ['Implement RBAC', 'Parameterized Queries']
    }

    # EVIDENCE VAULT
    evidence_vault = {
        'Date': [today_str, today_str],
        'Timestamp': [now_ts, now_ts],
        'Evidence_ID': ['EV-001', 'EV-002'],
        'Artifact_Source': ['Nmap Scan Log', 'Wireshark PCAP'],
        'SHA-256_Hash': ['e3b0c442...', '5df6e0e2...'],
        'Analyst': ['First Last Name', 'First Last Name'],
        'Integrity_Status': ['Verified', 'Verified']
    }

    # --- 3. GENERATION & STYLING ---
    try:
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            pd.DataFrame(dossier_data).to_excel(writer, sheet_name='Dossier', index=False, startrow=1)
            pd.DataFrame(blue_team_data).to_excel(writer, sheet_name='Blue_Team_NIST_IR', index=False, startrow=1)
            pd.DataFrame(red_team_data).to_excel(writer, sheet_name='Red_Team_PTES_OSSTMM', index=False, startrow=1)
            pd.DataFrame(owasp_data).to_excel(writer, sheet_name='Red_Team_OWASP', index=False, startrow=1)
            pd.DataFrame(evidence_vault).to_excel(writer, sheet_name='Evidence_Vault', index=False, startrow=1)

            styling_config = [
                ('Dossier', "INVESTIGATION DOSSIER SUMMARY", "212121", 2),
                ('Blue_Team_NIST_IR', "BLUE TEAM: TACTICAL INCIDENT RESPONSE (NIST ALIGNED)", "0D47A1", 14),
                ('Red_Team_PTES_OSSTMM', "RED TEAM: PTES & OSSTMM PENETRATION TESTING", "B71C1C", 11),
                ('Red_Team_OWASP', "RED TEAM: OWASP TOP 10 WEB ASSESSMENT", "B71C1C", 7),
                ('Evidence_Vault', "DFIR: EVIDENCE VAULT & HASH INTEGRITY", "424242", 7)
            ]

            for sheet, title, color, cols in styling_config:
                ws = writer.sheets[sheet]
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
                header = ws.cell(row=1, column=1)
                header.value = title
                header.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                header.font = Font(color="FFFFFF", bold=True, size=12)
                header.alignment = Alignment(horizontal="center")

                # Column Headers Row 2: Bold and Centered
                for col in range(1, cols + 1):
                    col_header = ws.cell(row=2, column=col)
                    col_header.font = Font(bold=True)
                    col_header.alignment = Alignment(horizontal="center")

        print("SUCCESS")
        os.startfile(filename)

    except Exception as e:
        print(f"SYSTEM ERROR: {e}")


if __name__ == "__main__":
    generate_ultimate_master_workbook()
